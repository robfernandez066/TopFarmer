from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10, color="000000")
GREEN = Font(name="Arial", size=10, color="008000")
HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
TITLE = Font(name="Arial", size=12, bold=True)
NOTE = Font(name="Arial", size=9, italic=True, color="555555")
HDRFILL = PatternFill("solid", fgColor="3B5B4A")
YELL = PatternFill("solid", fgColor="FFFF00")
THIN = Border(bottom=Side(style="thin", color="BBBBBB"))

wb = Workbook()

def header(ws, row, cols, widths=None):
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = HDR
        cell.fill = HDRFILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 30
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

def title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = TITLE
    if sub:
        ws["A2"] = sub
        ws["A2"].font = NOTE

# ----------------------------------------------------------------- LEGEND
ws = wb.active
ws.title = "Legend"
title(ws, "TopFarmer - Balance Sheet v0 (skeleton)",
      "Source of truth for every player-facing number. Game code must read exported CSVs, never hardcode values.")
rows = [
    ("", ""),
    ("HOW TO USE", ""),
    ("Blue text", "Hardcoded input. These are the cells you edit."),
    ("Black text", "Formula. Do not edit - it will be overwritten by the next tuning pass."),
    ("Green text", "Link to another tab."),
    ("Yellow fill", "Key global lever. Changing these moves the whole economy."),
    ("", ""),
    ("TABS", ""),
    ("Globals", "Economy-wide levers. Time multiplier, reward multiplier, price rules."),
    ("Crops", "Every plantable crop. Derived gold/hr and xp/hr per plot = the difficulty curve."),
    ("Goods", "Every produced ingredient and final good, with recipes and margins."),
    ("Buildings", "Production buildings, costs, unlock levels, parallel slots."),
    ("Progression", "Farm levels 1-10: XP curve, plots, storage, what unlocks."),
    ("Orders", "Order rewards derived from the value of goods requested. Formula-driven, not hand-tuned."),
    ("Currencies", "Gold and Starlight sources and sinks."),
    ("", ""),
    ("SCOPE OF v0", ""),
    ("", "3 crops, 2 production buildings, 10 levels. Deliberately rough."),
    ("", "Spec called for 1 building; 2 are included so the Sunwheat > Flour > Bread chain is representable."),
    ("", "Numbers here are placeholders chosen to be plausible, NOT tuned. Tuning is a separate roadmap phase."),
    ("", ""),
    ("EXPORT RULE", ""),
    ("", "Each tab exports to /data/<tabname>.csv. Both this .xlsx and the CSVs are committed."),
    ("", "Coder Agent never edits /data/ directly - it requests changes via progress.md."),
    ("", ""),
    ("WORKED EXAMPLE", ""),
    ("", "First pass priced Flour at 14g. Input cost (3 Sunwheat @ 6g) was 18g, so milling DESTROYED 4g of value."),
    ("", "Raw selling beat crafting - a trap the player would find before you did. Flour is now 24g. Check every margin after any price edit."),
    ("", ""),
    ("TIME MODEL", ""),
    ("", "All durations are real-world seconds. Growth and production continue while the app is closed."),
]
r = 4
for a, b in rows:
    ws.cell(row=r, column=1, value=a).font = Font(name="Arial", size=10, bold=bool(a and not b))
    ws.cell(row=r, column=2, value=b).font = Font(name="Arial", size=10)
    r += 1
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 105

# ----------------------------------------------------------------- GLOBALS
ws = wb.create_sheet("Globals")
title(ws, "Global Economy Levers",
      "Every time gate and reward rate in the game scales off these. Required by requirement E5 (monetization-readiness).")
header(ws, 4, ["Key", "Value", "Unit", "Notes"], [30, 12, 14, 78])
g = [
    ("GROWTH_TIME_MULT", 1.0, "multiplier", "Scales all crop growth durations. Lower = faster game. The primary pacing lever."),
    ("PRODUCTION_TIME_MULT", 1.0, "multiplier", "Scales all building production durations."),
    ("SELL_PRICE_MULT", 1.0, "multiplier", "Scales all sell prices. Use to inflate/deflate the economy without touching every row."),
    ("XP_MULT", 1.0, "multiplier", "Scales all XP awards."),
    ("ORDER_GOLD_MULT", 1.6, "multiplier", "Order gold reward = value of goods requested x this. Above 1.0 so orders beat raw selling."),
    ("ORDER_XP_MULT", 0.25, "multiplier", "Order XP = goods value x this."),
    ("TRAIN_BONUS_MULT", 2.5, "multiplier", "Large train/airship deliveries pay this instead of ORDER_GOLD_MULT."),
    ("STARLIGHT_ORDER_CHANCE", 0.15, "probability", "Chance a completed standard order also awards 1 Starlight."),
    ("STARLIGHT_HARVEST_CHANCE", 0.01, "probability", "Chance a harvest drops 1 Starlight."),
    ("XP_CURVE_BASE", 60, "xp", "XP required for level 2. Feeds the Progression curve."),
    ("XP_CURVE_EXP", 1.85, "exponent", "Level curve steepness. XP to reach level N = BASE x (N-1)^EXP."),
    ("STARTING_GOLD", 250, "gold", "New save gold."),
    ("STARTING_PLOTS", 4, "plots", "New save plot count."),
]
r = 5
for k, v, u, n in g:
    ws.cell(row=r, column=1, value=k).font = BLACK
    c = ws.cell(row=r, column=2, value=v); c.font = BLUE; c.fill = YELL
    ws.cell(row=r, column=3, value=u).font = BLACK
    ws.cell(row=r, column=4, value=n).font = NOTE
    r += 1
ws.cell(row=r+1, column=1, value="All values above are placeholders set by initial design, not derived from playtest data.").font = NOTE

# named-ish references (we use direct cell refs below)
GM = "Globals!$B$5"    # growth time mult
PM = "Globals!$B$6"    # production time mult
SP = "Globals!$B$7"    # sell price mult
XM = "Globals!$B$8"    # xp mult
OG = "Globals!$B$9"    # order gold mult
OX = "Globals!$B$10"   # order xp mult
TB = "Globals!$B$11"   # train bonus
XB = "Globals!$B$14"   # xp curve base
XE = "Globals!$B$15"   # xp curve exp

# ----------------------------------------------------------------- CROPS
ws = wb.create_sheet("Crops")
title(ws, "Crops",
      "Gold/hr and XP/hr per plot are the real outputs here. Later crops should trend upward on both, but not monotonically - give the player a reason to keep growing early crops.")
cols = ["id", "name", "unlock_level", "seed_cost", "growth_sec_base", "growth_sec_eff",
        "yield_qty", "sell_price_base", "sell_price_eff", "xp_base", "xp_eff",
        "revenue_per_harvest", "profit_per_harvest", "gold_per_hour", "xp_per_hour"]
header(ws, 4, cols, [14, 16, 13, 10, 13, 13, 9, 12, 12, 9, 9, 13, 13, 12, 11])
crops = [
    ("sunwheat",     "Sunwheat",     1, 3,   120,  2, 6,  2),
    ("duskcorn",     "Duskcorn",     3, 12,  600,  2, 22, 6),
    ("ember_pepper", "Ember Pepper", 6, 40,  1800, 3, 55, 16),
]
r = 5
for cid, name, lvl, seed, gsec, yq, price, xp in crops:
    ws.cell(row=r, column=1, value=cid).font = BLUE
    ws.cell(row=r, column=2, value=name).font = BLUE
    ws.cell(row=r, column=3, value=lvl).font = BLUE
    ws.cell(row=r, column=4, value=seed).font = BLUE
    ws.cell(row=r, column=5, value=gsec).font = BLUE
    ws.cell(row=r, column=6, value=f"=E{r}*{GM}").font = GREEN
    ws.cell(row=r, column=7, value=yq).font = BLUE
    ws.cell(row=r, column=8, value=price).font = BLUE
    ws.cell(row=r, column=9, value=f"=H{r}*{SP}").font = GREEN
    ws.cell(row=r, column=10, value=xp).font = BLUE
    ws.cell(row=r, column=11, value=f"=J{r}*{XM}").font = GREEN
    ws.cell(row=r, column=12, value=f"=G{r}*I{r}").font = BLACK
    ws.cell(row=r, column=13, value=f"=L{r}-D{r}").font = BLACK
    ws.cell(row=r, column=14, value=f"=IFERROR(M{r}/(F{r}/3600),0)").font = BLACK
    ws.cell(row=r, column=15, value=f"=IFERROR(K{r}/(F{r}/3600),0)").font = BLACK
    for c in range(1, 16):
        ws.cell(row=r, column=c).border = THIN
    r += 1
ws.cell(row=r+1, column=1, value="Placeholder values. sell_price_base is the design number; the game reads sell_price_eff.").font = NOTE
ws.cell(row=r+2, column=1, value="Rule of thumb to preserve when tuning: longer growth = higher gold/hr, so idle-friendly crops stay worth planting.").font = NOTE

# ----------------------------------------------------------------- GOODS
ws = wb.create_sheet("Goods")
title(ws, "Goods (ingredients and final goods)",
      "input_cost values inputs at their sell price - the opportunity cost of not selling them raw. margin below zero means the recipe destroys value.")
cols = ["id", "name", "building_id", "unlock_level", "input1_id", "input1_qty", "input2_id", "input2_qty",
        "prod_sec_base", "prod_sec_eff", "output_qty", "sell_price_base", "sell_price_eff",
        "xp_base", "xp_eff", "input_cost", "revenue", "margin", "gold_per_hour", "xp_per_hour"]
header(ws, 4, cols, [13, 14, 12, 12, 13, 10, 13, 10, 12, 12, 10, 12, 12, 9, 9, 11, 11, 11, 12, 11])
goods = [
    ("flour",      "Flour",      "flourmill", 2, "sunwheat", 3, None, 0, 300,  1, 24, 4),
    ("bread",      "Bread",      "bakery",    4, "flour",    2, None, 0, 600,  1, 72, 12),
    ("ember_oil",  "Ember Oil",  "flourmill", 7, "ember_pepper", 3, None, 0, 900, 1, 210, 30),
    ("spice_loaf", "Spice Loaf", "bakery",    8, "bread", 1, "ember_oil", 1, 1200, 1, 420, 48),
]
CR_ID = "Crops!$A$5:$A$40"; CR_PR = "Crops!$I$5:$I$40"
GD_ID = "$A$5:$A$40"; GD_PR = "$M$5:$M$40"
r = 5
for gid, name, bld, lvl, i1, q1, i2, q2, psec, oq, price, xp in goods:
    ws.cell(row=r, column=1, value=gid).font = BLUE
    ws.cell(row=r, column=2, value=name).font = BLUE
    ws.cell(row=r, column=3, value=bld).font = BLUE
    ws.cell(row=r, column=4, value=lvl).font = BLUE
    ws.cell(row=r, column=5, value=i1).font = BLUE
    ws.cell(row=r, column=6, value=q1).font = BLUE
    ws.cell(row=r, column=7, value=i2 if i2 else "").font = BLUE
    ws.cell(row=r, column=8, value=q2).font = BLUE
    ws.cell(row=r, column=9, value=psec).font = BLUE
    ws.cell(row=r, column=10, value=f"=I{r}*{PM}").font = GREEN
    ws.cell(row=r, column=11, value=oq).font = BLUE
    ws.cell(row=r, column=12, value=price).font = BLUE
    ws.cell(row=r, column=13, value=f"=L{r}*{SP}").font = GREEN
    ws.cell(row=r, column=14, value=xp).font = BLUE
    ws.cell(row=r, column=15, value=f"=N{r}*{XM}").font = GREEN
    look1 = (f'IFERROR(IFERROR(INDEX({CR_PR},MATCH($E{r},{CR_ID},0)),'
             f'INDEX({GD_PR},MATCH($E{r},{GD_ID},0)))*$F{r},0)')
    look2 = (f'IFERROR(IFERROR(INDEX({CR_PR},MATCH($G{r},{CR_ID},0)),'
             f'INDEX({GD_PR},MATCH($G{r},{GD_ID},0)))*$H{r},0)')
    ws.cell(row=r, column=16, value=f"={look1}+{look2}").font = BLACK
    ws.cell(row=r, column=17, value=f"=K{r}*M{r}").font = BLACK
    ws.cell(row=r, column=18, value=f"=Q{r}-P{r}").font = BLACK
    ws.cell(row=r, column=19, value=f"=IFERROR(R{r}/(J{r}/3600),0)").font = BLACK
    ws.cell(row=r, column=20, value=f"=IFERROR(O{r}/(J{r}/3600),0)").font = BLACK
    for c in range(1, 21):
        ws.cell(row=r, column=c).border = THIN
    r += 1
ws.cell(row=r+1, column=1, value="Leave input2_id blank for single-input recipes. Lookups search Crops first, then Goods.").font = NOTE
ws.cell(row=r+2, column=1, value="Placeholder values. Every margin must stay positive or the crafting chain is a trap.").font = NOTE

# ----------------------------------------------------------------- BUILDINGS
ws = wb.create_sheet("Buildings")
title(ws, "Production Buildings", "slots = how many jobs can run in parallel. A cheap early lever for both pacing and future monetization.")
header(ws, 4, ["id", "name", "unlock_level", "gold_cost", "build_sec", "slots", "notes"],
       [14, 16, 13, 11, 11, 8, 60])
blds = [
    ("flourmill", "Flourmill", 2, 300, 300, 2, "First production building. Unlocks the Sunwheat chain."),
    ("bakery",    "Bakery",    4, 1200, 900, 2, "Second tier. Consumes Flour."),
]
r = 5
for bid, name, lvl, cost, bsec, slots, note in blds:
    for i, v in enumerate([bid, name, lvl, cost, bsec, slots], start=1):
        ws.cell(row=r, column=i, value=v).font = BLUE
    ws.cell(row=r, column=7, value=note).font = NOTE
    r += 1
ws.cell(row=r+1, column=1, value="Additional slots per building are a natural later paid upgrade (requirement E5). Keep slots data-driven.").font = NOTE

# ----------------------------------------------------------------- PROGRESSION
ws = wb.create_sheet("Progression")
title(ws, "Farm Levels 1-10", "XP to reach level N = XP_CURVE_BASE x (N-1)^XP_CURVE_EXP. Change the curve in Globals, not here.")
header(ws, 4, ["level", "xp_to_reach", "cumulative_xp", "plots_unlocked", "storage_cap", "unlocks"],
       [8, 13, 14, 14, 12, 62])
unlocks = {
    1: "Sunwheat. Starting plots.", 2: "Flourmill. Flour.", 3: "Duskcorn. Order board.",
    4: "Bakery. Bread.", 5: "Territory expansion I. Decor storage.",
    6: "Ember Pepper. Pets.", 7: "Ember Oil. Daily quests.",
    8: "Spice Loaf. Train deliveries.", 9: "Territory expansion II.",
    10: "Weekly quests. Layout saving.",
}
r = 5
for lvl in range(1, 11):
    ws.cell(row=r, column=1, value=lvl).font = BLUE
    if lvl == 1:
        ws.cell(row=r, column=2, value=0).font = BLACK
        ws.cell(row=r, column=3, value=0).font = BLACK
    else:
        ws.cell(row=r, column=2, value=f"=ROUND({XB}*(A{r}-1)^{XE},0)").font = GREEN
        ws.cell(row=r, column=3, value=f"=C{r-1}+B{r}").font = BLACK
    ws.cell(row=r, column=4, value=f"=Globals!$B$17+INT((A{r}-1)*1.5)").font = GREEN
    ws.cell(row=r, column=5, value=f"=50+(A{r}-1)*25").font = BLACK
    ws.cell(row=r, column=6, value=unlocks[lvl]).font = BLUE
    r += 1
ws.cell(row=r+1, column=1, value="plots and storage_cap use placeholder formulas, not design intent. Replace with a considered curve in the tuning phase.").font = NOTE

# ----------------------------------------------------------------- ORDERS
ws = wb.create_sheet("Orders")
title(ws, "Orders and Deliveries",
      "Rewards are derived from the value of what is requested. Never hand-write a gold reward - it guarantees drift.")
header(ws, 4, ["order_id", "type", "min_level", "item1_id", "qty1", "item2_id", "qty2", "item3_id", "qty3",
               "goods_value", "gold_reward", "xp_reward", "starlight_reward"],
       [12, 12, 11, 14, 8, 14, 8, 14, 8, 12, 12, 11, 14])
orders = [
    ("ORD_S_01", "standard", 3, "sunwheat", 8, "", 0, "", 0),
    ("ORD_S_02", "standard", 4, "flour", 4, "sunwheat", 6, "", 0),
    ("ORD_S_03", "standard", 5, "bread", 3, "flour", 2, "", 0),
    ("ORD_T_01", "train",    8, "bread", 10, "ember_oil", 4, "spice_loaf", 2),
]
def val(col, qty_col, r):
    return (f'IFERROR(IFERROR(INDEX({CR_PR},MATCH(${col}{r},{CR_ID},0)),'
            f'INDEX(Goods!$M$5:$M$40,MATCH(${col}{r},Goods!$A$5:$A$40,0)))*${qty_col}{r},0)')
r = 5
for oid, typ, lvl, i1, q1, i2, q2, i3, q3 in orders:
    for i, v in enumerate([oid, typ, lvl, i1, q1, i2, q2, i3, q3], start=1):
        ws.cell(row=r, column=i, value=v).font = BLUE
    ws.cell(row=r, column=10, value=f"={val('D','E',r)}+{val('F','G',r)}+{val('H','I',r)}").font = BLACK
    ws.cell(row=r, column=11,
            value=f'=ROUND(J{r}*IF($B{r}="train",{TB},{OG}),0)').font = GREEN
    ws.cell(row=r, column=12, value=f"=ROUND(J{r}*{OX},0)").font = GREEN
    ws.cell(row=r, column=13, value=f'=IF($B{r}="train",2,0)').font = BLACK
    r += 1
ws.cell(row=r+1, column=1, value="These four are format examples. Live orders should be generated at runtime from unlocked items using these same formulas.").font = NOTE
ws.cell(row=r+2, column=1, value="Standard orders award Starlight probabilistically (Globals STARLIGHT_ORDER_CHANCE); train orders award it deterministically.").font = NOTE

# ----------------------------------------------------------------- CURRENCIES
ws = wb.create_sheet("Currencies")
title(ws, "Currencies", "Gold is the working currency. Starlight is scarce and must always have more demand than supply.")
header(ws, 4, ["currency", "flow", "source_or_sink", "amount", "notes"], [12, 8, 30, 12, 62])
cur = [
    ("gold", "source", "Selling crops", "see Crops", "Baseline income. Deliberately the worst gold/hr option."),
    ("gold", "source", "Selling goods", "see Goods", "Better than raw crops after accounting for time."),
    ("gold", "source", "Standard orders", "value x ORDER_GOLD_MULT", "Best routine income. Drives engagement with the order board."),
    ("gold", "source", "Train deliveries", "value x TRAIN_BONUS_MULT", "Large, infrequent, long completion window."),
    ("gold", "sink", "Seeds", "see Crops", "Small continuous drain."),
    ("gold", "sink", "Buildings", "see Buildings", "Major early sink."),
    ("gold", "sink", "Territory expansion", "TBD", "Should be the dominant mid-game sink. Not yet specified."),
    ("gold", "sink", "Decor", "TBD", "Optional sink. Needed or late-game gold becomes meaningless."),
    ("starlight", "source", "Order completion", "1 @ 15% chance", "Globals STARLIGHT_ORDER_CHANCE."),
    ("starlight", "source", "Harvest drop", "1 @ 1% chance", "Globals STARLIGHT_HARVEST_CHANCE."),
    ("starlight", "source", "Train delivery", "2 guaranteed", "The reliable path. Rewards planning."),
    ("starlight", "sink", "Growth speedup", "TBD", "Reserved as a plausible future paid feature (E5). Keep earnable in v1."),
    ("starlight", "sink", "Extra plot slot", "TBD", "Same."),
    ("starlight", "sink", "Storage upgrade", "TBD", "Same."),
]
r = 5
for c1, c2, c3, c4, c5 in cur:
    for i, v in enumerate([c1, c2, c3, c4], start=1):
        ws.cell(row=r, column=i, value=v).font = BLUE
    ws.cell(row=r, column=5, value=c5).font = NOTE
    r += 1
ws.cell(row=r+1, column=1, value="TBD rows are unresolved design questions, not omissions. They must be closed before the tuning phase.").font = NOTE

for s in wb.worksheets:
    s.sheet_view.showGridLines = False
    s.freeze_panes = "A5"

wb.save("/home/claude/balance.xlsx")
print("ok")
